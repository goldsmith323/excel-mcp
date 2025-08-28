"""
Engineering-focused Excel analysis tools for MCP server
Specialized for engineering calculators, formulas, and technical documentation
"""

import re
import json
import logging
from typing import Dict, List, Tuple, Any, Optional
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np


class EngineeringExcelAnalyzer:
    """Advanced analyzer for engineering Excel calculators"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = load_workbook(file_path, data_only=False)
        self.workbook_values = load_workbook(file_path, data_only=True)
        
        # Engineering patterns
        self.unit_patterns = {
            'length': r'\b(mm|cm|m|km|in|ft|yd|mil)\b',
            'area': r'\b(mm²|cm²|m²|km²|in²|ft²|yd²|sqft|sqin)\b',
            'volume': r'\b(mm³|cm³|m³|L|gal|ft³|in³|cuft|cuin)\b',
            'mass': r'\b(g|kg|lb|lbs|ton|tonnes|oz)\b',
            'force': r'\b(N|kN|MN|lbf|kip|kips)\b',
            'pressure': r'\b(Pa|kPa|MPa|GPa|psi|psf|bar|atm)\b',
            'time': r'\b(s|sec|min|hr|hour|ms|msec|μs)\b',
            'temperature': r'\b(°C|°F|K|°R)\b',
            'angle': r'\b(deg|rad|°|degrees|radians)\b',
            'frequency': r'\b(Hz|kHz|MHz|GHz|rpm)\b'
        }
        
        self.engineering_keywords = {
            'inputs': ['input', 'parameter', 'given', 'data', 'enter', 'specify'],
            'outputs': ['output', 'result', 'calculated', 'answer', 'solution'],
            'calculations': ['formula', 'equation', 'calculate', 'compute', 'solve'],
            'constants': ['constant', 'factor', 'coefficient', 'property'],
            'validation': ['check', 'verify', 'validate', 'limit', 'maximum', 'minimum'],
            'documentation': ['description', 'notes', 'reference', 'source', 'standard', 'code']
        }
        
    def analyze_calculator_structure(self) -> Dict[str, Any]:
        """Comprehensive analysis of engineering calculator structure"""
        try:
            analysis = {
                'calculator_info': self._identify_calculator_type(),
                'sheet_analysis': self._analyze_all_sheets(),
                'input_parameters': self._find_input_parameters(),
                'output_parameters': self._find_output_parameters(),
                'formulas': self._analyze_formulas(),
                'units_analysis': self._analyze_units(),
                'validation_rules': self._find_validation_rules(),
                'documentation': self._extract_documentation(),
                'dependencies': self._analyze_dependencies(),
                'engineering_standards': self._identify_standards()
            }
            
            return analysis
            
        except Exception as e:
            logging.error(f"Error analyzing calculator structure: {e}")
            return {'error': str(e)}
    
    def _identify_calculator_type(self) -> Dict[str, Any]:
        """Identify the type and purpose of the engineering calculator"""
        calculator_info = {
            'file_name': self.file_path.split('/')[-1],
            'sheet_names': self.workbook.sheetnames,
            'total_sheets': len(self.workbook.sheetnames),
            'calculator_type': 'Unknown',
            'engineering_domain': 'Unknown',
            'purpose': 'Unknown'
        }
        
        # Analyze sheet names and content for calculator type
        sheet_text = ' '.join(self.workbook.sheetnames).lower()
        
        # Engineering domain detection
        if any(word in sheet_text for word in ['blast', 'explosion', 'pressure', 'ufc']):
            calculator_info['engineering_domain'] = 'Blast/Explosive Engineering'
            calculator_info['calculator_type'] = 'Blast Load Calculator'
            calculator_info['purpose'] = 'Calculate blast pressures and structural loads'
        elif any(word in sheet_text for word in ['beam', 'column', 'structural', 'load']):
            calculator_info['engineering_domain'] = 'Structural Engineering'
        elif any(word in sheet_text for word in ['thermal', 'heat', 'temp']):
            calculator_info['engineering_domain'] = 'Thermal Engineering'
        elif any(word in sheet_text for word in ['fluid', 'flow', 'pipe']):
            calculator_info['engineering_domain'] = 'Fluid Mechanics'
        elif any(word in sheet_text for word in ['electrical', 'voltage', 'current']):
            calculator_info['engineering_domain'] = 'Electrical Engineering'
        
        return calculator_info
    
    def _analyze_all_sheets(self) -> Dict[str, Dict[str, Any]]:
        """Analyze each sheet's purpose and content"""
        sheet_analysis = {}
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_values = self.workbook_values[sheet_name]
            
            # Basic sheet info
            sheet_info = {
                'name': sheet_name,
                'max_row': sheet.max_row,
                'max_column': sheet.max_column,
                'has_formulas': False,
                'has_data': False,
                'sheet_type': 'Unknown',
                'purpose': 'Unknown'
            }
            
            # Analyze sheet content
            formulas = []
            data_cells = 0
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        data_cells += 1
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            formulas.append(cell.value)
                            sheet_info['has_formulas'] = True
            
            sheet_info['has_data'] = data_cells > 0
            sheet_info['formula_count'] = len(formulas)
            sheet_info['data_cell_count'] = data_cells
            
            # Determine sheet type based on content
            sheet_name_lower = sheet_name.lower()
            if 'setup' in sheet_name_lower or 'input' in sheet_name_lower:
                sheet_info['sheet_type'] = 'Input/Configuration'
            elif 'output' in sheet_name_lower or 'result' in sheet_name_lower:
                sheet_info['sheet_type'] = 'Output/Results'
            elif 'calc' in sheet_name_lower or 'computation' in sheet_name_lower:
                sheet_info['sheet_type'] = 'Calculation'
            elif len(formulas) > 20:
                sheet_info['sheet_type'] = 'Calculation'
            elif 'lookup' in sheet_name_lower or 'table' in sheet_name_lower:
                sheet_info['sheet_type'] = 'Lookup Table'
            elif data_cells > 50 and len(formulas) < 5:
                sheet_info['sheet_type'] = 'Data/Constants'
            
            sheet_analysis[sheet_name] = sheet_info
        
        return sheet_analysis
    
    def _find_input_parameters(self) -> List[Dict[str, Any]]:
        """Find and analyze input parameters across all sheets"""
        input_params = []
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_values = self.workbook_values[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_text = cell.value.lower()
                        
                        # Look for input indicators
                        if any(keyword in cell_text for keyword in self.engineering_keywords['inputs']):
                            # Check adjacent cells for values and units
                            param_info = self._analyze_parameter_cell(sheet, sheet_values, cell)
                            if param_info:
                                param_info['sheet'] = sheet_name
                                param_info['type'] = 'input'
                                input_params.append(param_info)
        
        return input_params
    
    def _find_output_parameters(self) -> List[Dict[str, Any]]:
        """Find and analyze output parameters"""
        output_params = []
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_values = self.workbook_values[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_text = cell.value.lower()
                        
                        # Look for output indicators
                        if any(keyword in cell_text for keyword in self.engineering_keywords['outputs']):
                            param_info = self._analyze_parameter_cell(sheet, sheet_values, cell)
                            if param_info:
                                param_info['sheet'] = sheet_name
                                param_info['type'] = 'output'
                                output_params.append(param_info)
        
        return output_params
    
    def _analyze_parameter_cell(self, sheet, sheet_values, cell) -> Optional[Dict[str, Any]]:
        """Analyze a specific parameter cell and its surroundings"""
        try:
            row, col = cell.row, cell.column
            
            param_info = {
                'name': cell.value,
                'location': f"{get_column_letter(col)}{row}",
                'row': row,
                'column': col,
                'value': None,
                'units': None,
                'description': None,
                'formula': None
            }
            
            # Look for value in adjacent cells
            for dr, dc in [(0, 1), (0, 2), (1, 0), (0, -1), (-1, 0)]:
                try:
                    adjacent_row, adjacent_col = row + dr, col + dc
                    if adjacent_row > 0 and adjacent_col > 0:
                        value_cell = sheet.cell(adjacent_row, adjacent_col)
                        value_cell_values = sheet_values.cell(adjacent_row, adjacent_col)
                        
                        if value_cell.value and str(value_cell.value).startswith('='):
                            param_info['formula'] = value_cell.value
                            param_info['value'] = value_cell_values.value
                        elif value_cell_values.value is not None:
                            if isinstance(value_cell_values.value, (int, float)):
                                param_info['value'] = value_cell_values.value
                except:
                    continue
            
            # Look for units
            param_info['units'] = self._extract_units_from_text(cell.value)
            
            return param_info if param_info['value'] is not None else None
            
        except Exception as e:
            logging.error(f"Error analyzing parameter cell: {e}")
            return None
    
    def _analyze_formulas(self) -> Dict[str, List[Dict[str, Any]]]:
        """Analyze all formulas in the workbook"""
        formulas_by_sheet = {}
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_formulas = []
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_info = {
                            'location': f"{get_column_letter(cell.column)}{cell.row}",
                            'formula': cell.value,
                            'complexity': self._assess_formula_complexity(cell.value),
                            'functions_used': self._extract_excel_functions(cell.value),
                            'references': self._extract_cell_references(cell.value)
                        }
                        sheet_formulas.append(formula_info)
            
            formulas_by_sheet[sheet_name] = sheet_formulas
        
        return formulas_by_sheet
    
    def _analyze_units(self) -> Dict[str, List[str]]:
        """Analyze units used throughout the calculator"""
        units_found = {}
        
        for unit_type, pattern in self.unit_patterns.items():
            units_found[unit_type] = []
            
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            matches = re.findall(pattern, cell.value, re.IGNORECASE)
                            units_found[unit_type].extend(matches)
            
            # Remove duplicates and sort
            units_found[unit_type] = sorted(list(set(units_found[unit_type])))
        
        return units_found
    
    def _find_validation_rules(self) -> List[Dict[str, Any]]:
        """Find validation rules and constraints"""
        validation_rules = []
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_text = cell.value.lower()
                        
                        # Look for validation keywords
                        if any(keyword in cell_text for keyword in self.engineering_keywords['validation']):
                            rule_info = {
                                'sheet': sheet_name,
                                'location': f"{get_column_letter(cell.column)}{cell.row}",
                                'rule_text': cell.value,
                                'type': 'constraint'
                            }
                            validation_rules.append(rule_info)
        
        return validation_rules
    
    def _extract_documentation(self) -> Dict[str, List[str]]:
        """Extract documentation, notes, and references"""
        documentation = {
            'descriptions': [],
            'notes': [],
            'references': [],
            'standards': []
        }
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and len(cell.value) > 20:
                        cell_text = cell.value.lower()
                        
                        if any(keyword in cell_text for keyword in self.engineering_keywords['documentation']):
                            if 'reference' in cell_text or 'ref' in cell_text:
                                documentation['references'].append(cell.value)
                            elif 'note' in cell_text:
                                documentation['notes'].append(cell.value)
                            elif 'standard' in cell_text or 'code' in cell_text:
                                documentation['standards'].append(cell.value)
                            else:
                                documentation['descriptions'].append(cell.value)
        
        return documentation
    
    def _analyze_dependencies(self) -> Dict[str, List[str]]:
        """Analyze dependencies between sheets and cells"""
        dependencies = {}
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_deps = []
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # Extract external sheet references
                        external_refs = re.findall(r"'([^']+)'!", cell.value)
                        sheet_deps.extend(external_refs)
            
            dependencies[sheet_name] = list(set(sheet_deps))
        
        return dependencies
    
    def _identify_standards(self) -> List[str]:
        """Identify engineering standards and codes referenced"""
        standards = []
        standard_patterns = [
            r'\bUFC\s*\d+', r'\bAISC\s*\d+', r'\bASCE\s*\d+', r'\bACI\s*\d+',
            r'\bIBC\s*\d+', r'\bAPI\s*\d+', r'\bASTM\s*[A-Z]\d+', r'\bISO\s*\d+',
            r'\bEN\s*\d+', r'\bBS\s*\d+', r'\bAWS\s*[A-Z]\d+'
        ]
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        for pattern in standard_patterns:
                            matches = re.findall(pattern, cell.value, re.IGNORECASE)
                            standards.extend(matches)
        
        return list(set(standards))
    
    def _extract_units_from_text(self, text: str) -> Optional[str]:
        """Extract units from text string"""
        if not isinstance(text, str):
            return None
            
        for unit_type, pattern in self.unit_patterns.items():
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                return matches[0]
        return None
    
    def _assess_formula_complexity(self, formula: str) -> str:
        """Assess the complexity of a formula"""
        if not formula:
            return 'simple'
        
        complexity_indicators = [
            ('IF', 'nested_if', 2),
            ('LOOKUP', 'lookup', 3),
            ('INDEX', 'array', 3),
            ('SUMPRODUCT', 'array', 3),
            ('VLOOKUP', 'lookup', 2),
            ('HLOOKUP', 'lookup', 2)
        ]
        
        complexity_score = 0
        for func, category, score in complexity_indicators:
            complexity_score += formula.upper().count(func) * score
        
        if complexity_score > 10:
            return 'very_complex'
        elif complexity_score > 5:
            return 'complex' 
        elif complexity_score > 2:
            return 'moderate'
        else:
            return 'simple'
    
    def _extract_excel_functions(self, formula: str) -> List[str]:
        """Extract Excel functions used in formula"""
        if not formula:
            return []
        
        # Common Excel functions
        functions = re.findall(r'\b([A-Z]+)\s*\(', formula.upper())
        return list(set(functions))
    
    def _extract_cell_references(self, formula: str) -> List[str]:
        """Extract cell references from formula"""
        if not formula:
            return []
        
        # Cell references pattern (e.g., A1, $B$2, Sheet1!A1)
        refs = re.findall(r"(?:'[^']+\'!)?(?:\$?[A-Z]+\$?\d+)", formula.upper())
        return list(set(refs))
    
    def get_calculation_summary(self) -> Dict[str, Any]:
        """Get a high-level summary of calculator capabilities"""
        analysis = self.analyze_calculator_structure()
        
        summary = {
            'calculator_type': analysis['calculator_info']['calculator_type'],
            'engineering_domain': analysis['calculator_info']['engineering_domain'],
            'total_inputs': len(analysis['input_parameters']),
            'total_outputs': len(analysis['output_parameters']),
            'total_formulas': sum(len(formulas) for formulas in analysis['formulas'].values()),
            'units_used': [unit for unit_list in analysis['units_analysis'].values() for unit in unit_list],
            'standards_referenced': analysis['engineering_standards'],
            'sheet_types': {sheet: info['sheet_type'] for sheet, info in analysis['sheet_analysis'].items()}
        }
        
        return summary