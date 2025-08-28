"""Excel operations module."""

import os
import glob
from typing import Dict, Any, List, Optional
from openpyxl import load_workbook
from openpyxl.workbook import Workbook


class ExcelHandler:
    """Handles Excel file operations."""
    
    def __init__(self, folder_path: str):
        """Initialize with Excel folder path."""
        self.folder_path = folder_path
        self.current_file: Optional[str] = None
        if not os.path.exists(folder_path):
            raise FileNotFoundError(f"Excel folder not found: {folder_path}")
    
    def list_excel_files(self) -> Dict[str, Any]:
        """List all Excel files in the folder."""
        pattern = os.path.join(self.folder_path, "*.xlsx")
        excel_files = glob.glob(pattern)
        
        files_info = []
        for file_path in excel_files:
            filename = os.path.basename(file_path)
            file_size = os.path.getsize(file_path)
            
            # Try to get basic info about sheets
            try:
                workbook = load_workbook(file_path)
                sheet_count = len(workbook.sheetnames)
                sheet_names = workbook.sheetnames[:3]  # First 3 sheets
                if len(workbook.sheetnames) > 3:
                    sheet_names.append(f"... and {len(workbook.sheetnames) - 3} more")
            except Exception:
                sheet_count = "Unknown"
                sheet_names = ["Unable to read"]
            
            files_info.append({
                "filename": filename,
                "full_path": file_path,
                "size_bytes": file_size,
                "size_mb": round(file_size / (1024*1024), 2),
                "sheet_count": sheet_count,
                "sheet_preview": sheet_names
            })
        
        return {
            "folder_path": self.folder_path,
            "file_count": len(files_info),
            "files": files_info,
            "current_selected": self.current_file
        }
    
    def select_excel_file(self, filename: str) -> Dict[str, Any]:
        """Select an Excel file to work with."""
        file_path = os.path.join(self.folder_path, filename)
        
        if not os.path.exists(file_path):
            available_files = [os.path.basename(f) for f in glob.glob(os.path.join(self.folder_path, "*.xlsx"))]
            raise FileNotFoundError(f"Excel file '{filename}' not found. Available files: {available_files}")
        
        self.current_file = file_path
        
        # Get basic info about the selected file
        try:
            workbook = load_workbook(file_path)
            sheet_info = [{
                "name": sheet_name,
                "max_row": workbook[sheet_name].max_row,
                "max_column": workbook[sheet_name].max_column
            } for sheet_name in workbook.sheetnames]
        except Exception as e:
            sheet_info = []
        
        return {
            "selected_file": filename,
            "full_path": file_path,
            "file_size": os.path.getsize(file_path),
            "sheet_count": len(sheet_info),
            "sheets": sheet_info,
            "status": "Selected successfully"
        }
    
    def find_excel_files_by_keyword(self, keyword: str) -> List[Dict[str, Any]]:
        """Find Excel files that match a keyword in their filename."""
        all_files = self.list_excel_files()["files"]
        
        matching_files = []
        keyword_lower = keyword.lower()
        
        for file_info in all_files:
            filename_lower = file_info["filename"].lower()
            if keyword_lower in filename_lower:
                # Calculate relevance score based on how well the keyword matches
                if filename_lower.startswith(keyword_lower):
                    score = 100  # Perfect prefix match
                elif keyword_lower in filename_lower.split("_")[0]:
                    score = 80   # Matches first part of filename
                else:
                    score = 50   # Contains keyword somewhere
                
                file_info["relevance_score"] = score
                matching_files.append(file_info)
        
        # Sort by relevance score (highest first)
        matching_files.sort(key=lambda x: x["relevance_score"], reverse=True)
        
        return matching_files
    
    def get_document_info(self) -> Dict[str, Any]:
        """Get basic information about the currently selected Excel document."""
        if not self.current_file:
            raise ValueError("No Excel file selected. Use select_excel_file() first.")
        
        workbook = load_workbook(self.current_file)
        
        sheet_info = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            sheet_info.append({
                "name": sheet_name,
                "max_row": max_row,
                "max_column": max_col,
                "dimensions": f"{max_row}x{max_col}"
            })
        
        return {
            "file_path": self.current_file,
            "filename": os.path.basename(self.current_file),
            "file_size": os.path.getsize(self.current_file),
            "sheet_count": len(workbook.sheetnames),
            "sheets": sheet_info
        }
    
    def update_cell(self, sheet_name: str, cell_address: str, value: Any) -> Dict[str, Any]:
        """Update a cell value in the specified sheet of the currently selected file."""
        if not self.current_file:
            raise ValueError("No Excel file selected. Use select_excel_file() first.")
        
        workbook = load_workbook(self.current_file)
        
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
        
        sheet = workbook[sheet_name]
        
        # Handle merged cells and regular cells
        try:
            cell = sheet[cell_address]
            # Check if this is a merged cell
            if hasattr(cell, 'coordinate') and any(cell.coordinate in merged_range for merged_range in sheet.merged_cells.ranges):
                # Find the top-left cell of the merged range
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        top_left = merged_range.start_cell
                        old_value = top_left.value
                        # Unmerge, update, and re-merge
                        sheet.unmerge_cells(str(merged_range))
                        sheet[cell_address] = value
                        sheet.merge_cells(str(merged_range))
                        break
            else:
                old_value = cell.value
                sheet[cell_address] = value
        except Exception as e:
            # If there's any issue with merged cells, try simple assignment
            try:
                old_value = sheet[cell_address].value
            except:
                old_value = "Unable to read"
            sheet[cell_address] = value
        
        # Save the workbook
        workbook.save(self.current_file)
        
        return {
            "file": os.path.basename(self.current_file),
            "sheet": sheet_name,
            "cell": cell_address,
            "old_value": old_value,
            "new_value": value,
            "success": True
        }