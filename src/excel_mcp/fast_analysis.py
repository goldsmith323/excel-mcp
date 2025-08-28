"""
Fast Excel Analysis - Optimized for speed and simple questions
Provides quick answers without heavy analysis
"""

import os
import re
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, Any, Optional


class FastExcelAnalyzer:
    """Lightweight, fast analyzer for quick Excel insights"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.file_name = os.path.basename(file_path)
        
        # Cache for basic info (load once, use multiple times)
        self._basic_info = None
        self._sheet_names = None
        
    def get_basic_info(self) -> Dict[str, Any]:
        """Get basic file info quickly - cached for performance"""
        if self._basic_info is not None:
            return self._basic_info
            
        try:
            # Load workbook once
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            
            self._sheet_names = wb.sheetnames
            self._basic_info = {
                'file_name': self.file_name,
                'file_path': self.file_path,
                'file_size_mb': round(os.path.getsize(self.file_path) / (1024*1024), 2),
                'sheet_count': len(wb.sheetnames),
                'sheet_names': wb.sheetnames
            }
            
            wb.close()
            return self._basic_info
            
        except Exception as e:
            return {'error': f"Cannot read file: {e}"}
    
    def quick_purpose_analysis(self) -> str:
        """Fast analysis of calculator purpose - under 2 seconds"""
        try:
            # Get basic info first
            info = self.get_basic_info()
            if 'error' in info:
                return f"❌ Error: {info['error']}"
            
            file_name = info['file_name'].lower()
            sheet_names = [name.lower() for name in info['sheet_names']]
            sheet_text = ' '.join(sheet_names + [file_name])
            
            # Quick pattern matching for common engineering calculators
            calculator_types = {
                'blast': ('🧨 Blast/Explosion Calculator', 'Calculates blast pressures and structural loads from explosives'),
                'beam': ('🏗️ Beam Analysis Calculator', 'Analyzes structural beam properties, loads, and deflections'),
                'column': ('🏛️ Column Design Calculator', 'Designs and analyzes structural columns'),
                'foundation': ('🏗️ Foundation Calculator', 'Calculates foundation loads and soil bearing capacity'),
                'thermal': ('🌡️ Thermal Analysis Calculator', 'Calculates heat transfer and thermal properties'),
                'fluid': ('💧 Fluid Mechanics Calculator', 'Analyzes fluid flow, pressure, and hydraulic systems'),
                'electrical': ('⚡ Electrical Calculator', 'Calculates electrical parameters like voltage, current, power'),
                'pressure': ('📊 Pressure Vessel Calculator', 'Analyzes pressure vessels and piping systems'),
                'wind': ('💨 Wind Load Calculator', 'Calculates wind loads on structures'),
                'seismic': ('🌍 Seismic Analysis Calculator', 'Analyzes earthquake loads and structural response'),
                'steel': ('🔩 Steel Design Calculator', 'Designs steel structural members'),
                'concrete': ('🧱 Concrete Design Calculator', 'Designs concrete structural elements'),
                'pipe': ('🔧 Piping Calculator', 'Calculates pipe sizing, pressure drops, flow rates'),
                'hvac': ('❄️ HVAC Calculator', 'Heating, ventilation, and air conditioning calculations'),
                'load': ('⚖️ Load Calculator', 'Calculates various structural loads')
            }
            
            # Find matches
            for keyword, (calc_type, description) in calculator_types.items():
                if keyword in sheet_text:
                    return f"📊 **Calculator Type**: {calc_type}\n\n📝 **Purpose**: {description}\n\n📁 **File**: {info['file_name']}\n📋 **Sheets**: {info['sheet_count']} ({', '.join(info['sheet_names'][:3])}{'...' if info['sheet_count'] > 3 else ''})"
            
            # If no specific match, provide generic analysis
            return f"📊 **Calculator Type**: Engineering Calculator\n\n📝 **Purpose**: General engineering calculations and analysis\n\n📁 **File**: {info['file_name']}\n📋 **Sheets**: {info['sheet_count']} ({', '.join(info['sheet_names'][:3])}{'...' if info['sheet_count'] > 3 else ''})"
            
        except Exception as e:
            return f"❌ Error analyzing calculator: {e}"
    
    def quick_summary(self) -> str:
        """Ultra-fast summary - under 1 second"""
        info = self.get_basic_info()
        if 'error' in info:
            return f"❌ {info['error']}"
        
        return f"📁 **{info['file_name']}**\n📋 {info['sheet_count']} sheets\n💾 {info['file_size_mb']} MB"
    
    def get_sheet_preview(self, sheet_name: str = None, max_rows: int = 5) -> str:
        """Quick preview of sheet data"""
        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            
            # Use first sheet if none specified
            if sheet_name is None:
                sheet_name = wb.sheetnames[0]
            
            if sheet_name not in wb.sheetnames:
                return f"❌ Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}"
            
            ws = wb[sheet_name]
            
            # Get first few rows of data
            preview = f"📋 **Sheet**: {sheet_name}\n\n"
            row_count = 0
            
            for row in ws.iter_rows(max_row=max_rows, values_only=True):
                if any(cell is not None for cell in row):
                    row_str = " | ".join([str(cell) if cell is not None else "" for cell in row[:6]])
                    preview += f"Row {row_count + 1}: {row_str}\n"
                    row_count += 1
                    if row_count >= max_rows:
                        break
            
            if row_count == 0:
                preview += "No data found in first few rows"
            
            wb.close()
            return preview
            
        except Exception as e:
            return f"❌ Error reading sheet: {e}"
    
    def find_key_values(self) -> str:
        """Find important values quickly by scanning for numbers"""
        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            key_values = []
            
            # Check first sheet only for speed
            ws = wb[wb.sheetnames[0]]
            
            # Look for cells with numbers that might be important
            for row in ws.iter_rows(max_row=20, max_col=10, values_only=True):
                for i, cell in enumerate(row):
                    if isinstance(cell, (int, float)) and cell != 0:
                        # Look for nearby text that might describe this value
                        try:
                            # Check cells to the left and above for labels
                            label_cell = ws.cell(ws.active.row - 1, i + 1).value if ws.active.row > 1 else None
                            if label_cell and isinstance(label_cell, str) and len(label_cell) < 50:
                                key_values.append(f"{label_cell}: {cell}")
                            elif len(key_values) < 5:  # Limit to prevent slowdown
                                key_values.append(f"Value: {cell}")
                        except:
                            continue
                        
                        if len(key_values) >= 5:  # Limit results for speed
                            break
                if len(key_values) >= 5:
                    break
            
            wb.close()
            
            if key_values:
                return f"🔢 **Key Values Found**:\n" + "\n".join([f"• {val}" for val in key_values[:5]])
            else:
                return "🔢 No obvious key values found in first sheet"
                
        except Exception as e:
            return f"❌ Error finding values: {e}"