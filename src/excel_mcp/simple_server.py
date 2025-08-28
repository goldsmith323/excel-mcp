#!/usr/bin/env python3
"""
Excel MCP Server - Simple Server for Engineering Calculator Analysis

This is the main MCP (Model Context Protocol) server that enables MCP clients
to interact with Excel files, particularly engineering calculators.

FEATURES:
- Basic Excel operations (read, write, update cells)
- Advanced engineering calculator analysis
- Input/output parameter identification
- Formula analysis and complexity assessment
- Units analysis and validation
- Engineering documentation extraction

USAGE:
    EXCEL_FILE_PATH="/path/to/file.xlsx" python3 simple_server.py

DEVELOPMENT:
    To extend this server:
    1. Add new tools in the @server.list_tools() function
    2. Add corresponding handlers in @server.call_tool() function
    3. Use engineering_tools.py for engineering-specific analysis
    4. Use excel_tools.py for basic Excel operations

DEPENDENCIES:
    - mcp: Model Context Protocol library
    - openpyxl: Excel file manipulation
    - pandas: Data analysis (used in engineering_tools.py)
    - numpy: Numerical operations (used in engineering_tools.py)
"""

import asyncio
import os
import sys
from mcp.server import Server
from mcp.server.models import InitializationOptions
import mcp.server.stdio
import mcp.types as types

try:
    from .excel_tools import ExcelHandler
    from .engineering_tools import EngineeringExcelAnalyzer
    from .fast_analysis import FastExcelAnalyzer
except ImportError:
    # When running directly (not as module)
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from excel_tools import ExcelHandler
    from engineering_tools import EngineeringExcelAnalyzer
    from fast_analysis import FastExcelAnalyzer


# ============================================================================
# MCP SERVER INITIALIZATION
# ============================================================================

# Initialize the MCP server with a unique name
server = Server("excel-mcp-simple")

# Global variables to track the connected Excel file
# These are set when the server starts with EXCEL_FILE_PATH environment variable
excel_handler: ExcelHandler = None  # Will be initialized if needed
current_file_path: str = None       # Path to the connected Excel file


# ============================================================================
# TOOL DEFINITIONS
# ============================================================================

@server.list_tools()
async def handle_list_tools() -> list[types.Tool]:
    """
    Define all available tools for Excel analysis.
    
    This function registers all tools that MCP clients can use to interact with Excel files.
    Tools are divided into two categories:
    1. Basic Excel Operations (get_document_info, get_sheet_data, update_cell, etc.)
    2. Engineering Analysis Tools (analyze_engineering_calculator, find_input_parameters, etc.)
    
    To add a new tool:
    1. Add a new types.Tool() entry here with name, description, and input schema
    2. Add a corresponding handler in handle_call_tool() function below
    """
    return [
        types.Tool(
            name="quick_purpose",
            description="FAST: Get the purpose and type of this calculator (under 2 seconds) - USE THIS FOR 'what is this calculator' questions",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        ),
        types.Tool(
            name="quick_summary",
            description="FAST: Get a quick summary of the Excel file (under 1 second) - USE THIS FOR basic file info",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        ),
        types.Tool(
            name="quick_preview",
            description="FAST: Preview sheet data quickly - USE THIS FOR 'show me the data' questions",
            inputSchema={
                "type": "object",
                "properties": {
                    "sheet_name": {
                        "type": "string",
                        "description": "Sheet name to preview (optional, uses first sheet if not specified)"
                    }
                },
                "required": []
            }
        ),
        types.Tool(
            name="get_document_info",
            description="DETAILED: Get comprehensive information about the current Excel document (slower)",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        ),
        types.Tool(
            name="get_sheet_data",
            description="Get data from a specific sheet with optional range",
            inputSchema={
                "type": "object",
                "properties": {
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the sheet to read"
                    },
                    "range": {
                        "type": "string",
                        "description": "Optional cell range (e.g., 'A1:C10'). If not specified, reads all data."
                    }
                },
                "required": ["sheet_name"]
            }
        ),
        types.Tool(
            name="update_cell",
            description="Update a single cell value",
            inputSchema={
                "type": "object",
                "properties": {
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the sheet to update"
                    },
                    "cell_address": {
                        "type": "string",
                        "description": "Cell address in A1 notation (e.g., 'A1', 'B5')"
                    },
                    "value": {
                        "description": "Value to set in the cell (string, number, or formula)"
                    }
                },
                "required": ["sheet_name", "cell_address", "value"]
            }
        ),
        types.Tool(
            name="update_range",
            description="Update multiple cells in a range",
            inputSchema={
                "type": "object",
                "properties": {
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the sheet to update"
                    },
                    "range": {
                        "type": "string",
                        "description": "Cell range in A1 notation (e.g., 'A1:C3')"
                    },
                    "values": {
                        "type": "array",
                        "description": "2D array of values to set (rows x columns)"
                    }
                },
                "required": ["sheet_name", "range", "values"]
            }
        ),
        types.Tool(
            name="add_sheet",
            description="Add a new worksheet to the Excel file",
            inputSchema={
                "type": "object",
                "properties": {
                    "sheet_name": {
                        "type": "string",
                        "description": "Name for the new sheet"
                    }
                },
                "required": ["sheet_name"]
            }
        ),
        types.Tool(
            name="analyze_engineering_calculator",
            description="Comprehensive analysis of engineering calculator structure, inputs, outputs, formulas, and documentation",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        ),
        types.Tool(
            name="get_calculator_summary",
            description="Get a high-level summary of the calculator's capabilities and engineering domain",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        ),
        types.Tool(
            name="find_input_parameters",
            description="Identify all input parameters with their locations, values, and units",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        ),
        types.Tool(
            name="find_output_parameters",
            description="Identify all output parameters with their locations, values, and units",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        ),
        types.Tool(
            name="analyze_formulas",
            description="Analyze all formulas in the calculator including complexity and dependencies",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        ),
        types.Tool(
            name="analyze_units",
            description="Analyze units used throughout the calculator and identify unit systems",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        ),
        types.Tool(
            name="extract_documentation",
            description="Extract engineering documentation, standards, references, and notes",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        ),
        types.Tool(
            name="validate_engineering_data",
            description="Find validation rules, constraints, and engineering limits",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        )
    ]


# ============================================================================
# TOOL HANDLERS
# ============================================================================

@server.call_tool()
async def handle_call_tool(name: str, arguments: dict) -> list[types.TextContent]:
    """
    Handle all tool execution requests from MCP clients.
    
    This function receives tool calls from MCP clients and routes them to
    the appropriate handlers. Each tool performs a specific Excel operation
    or engineering analysis.
    
    Args:
        name: The name of the tool to execute
        arguments: Dictionary of arguments passed to the tool
    
    Returns:
        List of TextContent responses to send back to Claude
    
    To add a new tool handler:
    1. Add an elif statement with your tool name
    2. Extract arguments using arguments.get("param_name")
    3. Perform the operation (use engineering_tools.py for analysis)
    4. Return formatted TextContent with results
    """
    if not current_file_path or not os.path.exists(current_file_path):
        return [types.TextContent(
            type="text",
            text="❌ No Excel file connected. Please use the Excel add-in to connect a file."
        )]
    
    try:
        if name == "quick_purpose":
            # FAST: Get calculator purpose in under 2 seconds
            fast_analyzer = FastExcelAnalyzer(current_file_path)
            result = fast_analyzer.quick_purpose_analysis()
            
            return [types.TextContent(type="text", text=result)]
        
        elif name == "quick_summary":
            # FAST: Get quick summary in under 1 second
            fast_analyzer = FastExcelAnalyzer(current_file_path)
            result = fast_analyzer.quick_summary()
            
            return [types.TextContent(type="text", text=result)]
        
        elif name == "quick_preview":
            # FAST: Preview sheet data
            fast_analyzer = FastExcelAnalyzer(current_file_path)
            sheet_name = arguments.get("sheet_name")
            result = fast_analyzer.get_sheet_preview(sheet_name)
            
            return [types.TextContent(type="text", text=result)]
            
        elif name == "get_document_info":
            # Create temporary single-file handler
            handler = SingleFileHandler(current_file_path)
            result = handler.get_document_info()
            
            return [types.TextContent(
                type="text",
                text=f"📊 Excel Document Information:\n\n"
                     f"📁 File: {result['filename']}\n"
                     f"🗂️ Path: {result['file_path']}\n"
                     f"💾 Size: {round(result['file_size']/(1024*1024), 2)} MB\n"
                     f"📋 Sheets: {result['sheet_count']}\n\n"
                     f"Sheet Details:\n" + 
                     "\n".join([f"  • {sheet['name']}: {sheet['dimensions']} "
                               f"({sheet['max_row']} rows × {sheet['max_column']} cols)" 
                               for sheet in result['sheets']])
            )]
            
        elif name == "get_sheet_data":
            handler = SingleFileHandler(current_file_path)
            sheet_name = arguments.get("sheet_name")
            range_spec = arguments.get("range")
            
            result = handler.get_sheet_data(sheet_name, range_spec)
            
            # Format the data for display
            data_preview = ""
            if result['data']:
                # Show first few rows
                preview_rows = result['data'][:5]
                for i, row in enumerate(preview_rows):
                    row_str = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    data_preview += f"Row {i+1}: {row_str}\n"
                
                if len(result['data']) > 5:
                    data_preview += f"... and {len(result['data']) - 5} more rows"
            else:
                data_preview = "No data found"
            
            return [types.TextContent(
                type="text",
                text=f"📊 Sheet Data: {sheet_name}\n\n"
                     f"📏 Range: {result['actual_range']}\n"
                     f"📈 Rows: {result['row_count']}, Columns: {result['col_count']}\n\n"
                     f"Data Preview:\n{data_preview}"
            )]
            
        elif name == "update_cell":
            handler = SingleFileHandler(current_file_path)
            sheet_name = arguments.get("sheet_name")
            cell_address = arguments.get("cell_address")
            value = arguments.get("value")
            
            result = handler.update_cell(sheet_name, cell_address, value)
            
            return [types.TextContent(
                type="text",
                text=f"✅ Cell Updated Successfully:\n\n"
                     f"📋 Sheet: {result['sheet']}\n"
                     f"📍 Cell: {result['cell']}\n"
                     f"🔄 Old value: {result['old_value']}\n"
                     f"✨ New value: {result['new_value']}"
            )]
            
        elif name == "update_range":
            handler = SingleFileHandler(current_file_path)
            sheet_name = arguments.get("sheet_name")
            range_spec = arguments.get("range")
            values = arguments.get("values")
            
            result = handler.update_range(sheet_name, range_spec, values)
            
            return [types.TextContent(
                type="text",
                text=f"✅ Range Updated Successfully:\n\n"
                     f"📋 Sheet: {result['sheet']}\n"
                     f"📏 Range: {result['range']}\n"
                     f"📊 Cells updated: {result['cells_updated']}\n"
                     f"✨ Operation completed"
            )]
            
        elif name == "add_sheet":
            handler = SingleFileHandler(current_file_path)
            sheet_name = arguments.get("sheet_name")
            
            result = handler.add_sheet(sheet_name)
            
            return [types.TextContent(
                type="text",
                text=f"✅ New Sheet Created:\n\n"
                     f"📋 Sheet name: {result['sheet_name']}\n"
                     f"📊 Total sheets now: {result['total_sheets']}\n"
                     f"✨ Ready for data input"
            )]
        
        elif name == "analyze_engineering_calculator":
            analyzer = EngineeringExcelAnalyzer(current_file_path)
            analysis = analyzer.analyze_calculator_structure()
            
            # Format comprehensive analysis
            calc_info = analysis['calculator_info']
            text = f"🔧 Engineering Calculator Analysis\n\n"
            text += f"📊 Calculator Type: {calc_info['calculator_type']}\n"
            text += f"🏗️  Engineering Domain: {calc_info['engineering_domain']}\n"
            text += f"📁 File: {calc_info['file_name']}\n"
            text += f"📋 Total Sheets: {calc_info['total_sheets']}\n\n"
            
            # Sheet analysis
            text += f"📄 Sheet Analysis:\n"
            for sheet_name, info in analysis['sheet_analysis'].items():
                text += f"  • {sheet_name}: {info['sheet_type']} ({info['data_cell_count']} cells, {info['formula_count']} formulas)\n"
            
            text += f"\n📊 Parameters:\n"
            text += f"  • Input Parameters: {len(analysis['input_parameters'])}\n"
            text += f"  • Output Parameters: {len(analysis['output_parameters'])}\n"
            text += f"  • Total Formulas: {sum(len(formulas) for formulas in analysis['formulas'].values())}\n"
            
            # Units
            units_summary = [unit for unit_list in analysis['units_analysis'].values() for unit in unit_list]
            text += f"  • Units Found: {len(units_summary)} ({', '.join(units_summary[:10])}{'...' if len(units_summary) > 10 else ''})\n"
            
            # Standards
            if analysis['engineering_standards']:
                text += f"\n📚 Engineering Standards: {', '.join(analysis['engineering_standards'])}\n"
            
            return [types.TextContent(type="text", text=text)]
        
        elif name == "get_calculator_summary":
            analyzer = EngineeringExcelAnalyzer(current_file_path)
            summary = analyzer.get_calculation_summary()
            
            text = f"📋 Calculator Summary\n\n"
            text += f"🔧 Type: {summary['calculator_type']}\n"
            text += f"🏗️  Domain: {summary['engineering_domain']}\n"
            text += f"📊 Inputs: {summary['total_inputs']}\n"
            text += f"📈 Outputs: {summary['total_outputs']}\n"
            text += f"🧮 Formulas: {summary['total_formulas']}\n"
            text += f"📏 Units: {', '.join(summary['units_used'][:15])}{'...' if len(summary['units_used']) > 15 else ''}\n"
            
            if summary['standards_referenced']:
                text += f"📚 Standards: {', '.join(summary['standards_referenced'])}\n"
            
            return [types.TextContent(type="text", text=text)]
        
        elif name == "find_input_parameters":
            analyzer = EngineeringExcelAnalyzer(current_file_path)
            analysis = analyzer.analyze_calculator_structure()
            inputs = analysis['input_parameters']
            
            if not inputs:
                text = "📊 No input parameters found with standard naming patterns."
            else:
                text = f"📊 Input Parameters Found ({len(inputs)}):\n\n"
                for i, param in enumerate(inputs[:20], 1):
                    text += f"{i}. {param['name']}\n"
                    text += f"   📍 Location: {param['sheet']}.{param['location']}\n"
                    if param['value'] is not None:
                        text += f"   💾 Value: {param['value']}\n"
                    if param['units']:
                        text += f"   📏 Units: {param['units']}\n"
                    if param['formula']:
                        text += f"   🧮 Formula: {param['formula'][:50]}...\n"
                    text += "\n"
                
                if len(inputs) > 20:
                    text += f"... and {len(inputs) - 20} more parameters\n"
            
            return [types.TextContent(type="text", text=text)]
        
        elif name == "find_output_parameters":
            analyzer = EngineeringExcelAnalyzer(current_file_path)
            analysis = analyzer.analyze_calculator_structure()
            outputs = analysis['output_parameters']
            
            if not outputs:
                text = "📈 No output parameters found with standard naming patterns."
            else:
                text = f"📈 Output Parameters Found ({len(outputs)}):\n\n"
                for i, param in enumerate(outputs[:20], 1):
                    text += f"{i}. {param['name']}\n"
                    text += f"   📍 Location: {param['sheet']}.{param['location']}\n"
                    if param['value'] is not None:
                        text += f"   💾 Value: {param['value']}\n"
                    if param['units']:
                        text += f"   📏 Units: {param['units']}\n"
                    if param['formula']:
                        text += f"   🧮 Formula: {param['formula'][:50]}...\n"
                    text += "\n"
                
                if len(outputs) > 20:
                    text += f"... and {len(outputs) - 20} more parameters\n"
            
            return [types.TextContent(type="text", text=text)]
        
        elif name == "analyze_formulas":
            analyzer = EngineeringExcelAnalyzer(current_file_path)
            analysis = analyzer.analyze_calculator_structure()
            formulas = analysis['formulas']
            
            text = f"🧮 Formula Analysis\n\n"
            total_formulas = sum(len(sheet_formulas) for sheet_formulas in formulas.values())
            text += f"📊 Total Formulas: {total_formulas}\n\n"
            
            for sheet_name, sheet_formulas in formulas.items():
                if sheet_formulas:
                    text += f"📋 {sheet_name} ({len(sheet_formulas)} formulas):\n"
                    
                    # Show complexity distribution
                    complexity_counts = {}
                    for formula in sheet_formulas:
                        comp = formula['complexity']
                        complexity_counts[comp] = complexity_counts.get(comp, 0) + 1
                    
                    for complexity, count in complexity_counts.items():
                        text += f"  • {complexity}: {count} formulas\n"
                    
                    # Show sample formulas
                    text += f"  Sample formulas:\n"
                    for i, formula in enumerate(sheet_formulas[:3]):
                        text += f"    {formula['location']}: {formula['formula'][:60]}...\n"
                    text += "\n"
            
            return [types.TextContent(type="text", text=text)]
        
        elif name == "analyze_units":
            analyzer = EngineeringExcelAnalyzer(current_file_path)
            analysis = analyzer.analyze_calculator_structure()
            units = analysis['units_analysis']
            
            text = f"📏 Units Analysis\n\n"
            
            for unit_type, unit_list in units.items():
                if unit_list:
                    text += f"📊 {unit_type.title()}: {', '.join(unit_list)}\n"
            
            # Determine unit system
            all_units = [unit for unit_list in units.values() for unit in unit_list]
            metric_units = ['mm', 'cm', 'm', 'km', 'kg', 'N', 'Pa', 'kPa', 'MPa']
            imperial_units = ['in', 'ft', 'yd', 'lb', 'lbf', 'psi', 'psf']
            
            metric_count = sum(1 for unit in all_units if unit in metric_units)
            imperial_count = sum(1 for unit in all_units if unit in imperial_units)
            
            text += f"\n🌐 Unit System Analysis:\n"
            text += f"  • Metric units: {metric_count}\n"
            text += f"  • Imperial units: {imperial_count}\n"
            
            if imperial_count > metric_count:
                text += f"  • Primary system: Imperial/US Customary\n"
            elif metric_count > imperial_count:
                text += f"  • Primary system: Metric/SI\n"
            else:
                text += f"  • Primary system: Mixed\n"
            
            return [types.TextContent(type="text", text=text)]
        
        elif name == "extract_documentation":
            analyzer = EngineeringExcelAnalyzer(current_file_path)
            analysis = analyzer.analyze_calculator_structure()
            docs = analysis['documentation']
            
            text = f"📚 Documentation Analysis\n\n"
            
            if docs['descriptions']:
                text += f"📝 Descriptions ({len(docs['descriptions'])}):\n"
                for desc in docs['descriptions'][:5]:
                    text += f"  • {desc[:100]}...\n"
                text += "\n"
            
            if docs['references']:
                text += f"📖 References ({len(docs['references'])}):\n"
                for ref in docs['references'][:5]:
                    text += f"  • {ref[:100]}...\n"
                text += "\n"
            
            if docs['standards']:
                text += f"📐 Standards ({len(docs['standards'])}):\n"
                for std in docs['standards'][:5]:
                    text += f"  • {std[:100]}...\n"
                text += "\n"
            
            if docs['notes']:
                text += f"📋 Notes ({len(docs['notes'])}):\n"
                for note in docs['notes'][:5]:
                    text += f"  • {note[:100]}...\n"
                text += "\n"
            
            # Engineering standards found
            if analysis['engineering_standards']:
                text += f"🏗️  Engineering Standards Referenced:\n"
                for std in analysis['engineering_standards']:
                    text += f"  • {std}\n"
            
            return [types.TextContent(type="text", text=text)]
        
        elif name == "validate_engineering_data":
            analyzer = EngineeringExcelAnalyzer(current_file_path)
            analysis = analyzer.analyze_calculator_structure()
            validation = analysis['validation_rules']
            
            if not validation:
                text = "🔍 No explicit validation rules found in standard patterns."
            else:
                text = f"🔍 Validation Rules Found ({len(validation)}):\n\n"
                for i, rule in enumerate(validation[:10], 1):
                    text += f"{i}. {rule['rule_text']}\n"
                    text += f"   📍 Location: {rule['sheet']}.{rule['location']}\n"
                    text += f"   🔧 Type: {rule['type']}\n\n"
                
                if len(validation) > 10:
                    text += f"... and {len(validation) - 10} more rules\n"
            
            # Add general engineering guidance
            text += f"\n💡 General Engineering Validation Recommendations:\n"
            text += f"  • Verify input ranges are within realistic engineering limits\n"
            text += f"  • Check unit consistency throughout calculations\n"
            text += f"  • Validate against applicable engineering standards\n"
            text += f"  • Confirm formulas match referenced design codes\n"
            
            return [types.TextContent(type="text", text=text)]
        
        else:
            return [types.TextContent(
                type="text",
                text=f"❌ Unknown tool: {name}"
            )]
    
    except Exception as e:
        return [types.TextContent(
            type="text",
            text=f"❌ Error executing {name}: {str(e)}"
        )]


# ============================================================================
# EXCEL FILE HANDLER
# ============================================================================

class SingleFileHandler:
    """
    Handler for basic Excel file operations.
    
    This class provides low-level Excel file operations like reading sheets,
    updating cells, and adding sheets. For advanced engineering analysis,
    use the EngineeringExcelAnalyzer class from engineering_tools.py.
    
    Args:
        file_path: Path to the Excel file to operate on
    
    Usage:
        handler = SingleFileHandler("/path/to/file.xlsx")
        info = handler.get_document_info()
        data = handler.get_sheet_data("Sheet1")
    """
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        
    def get_document_info(self):
        """Get document information."""
        from openpyxl import load_workbook
        
        workbook = load_workbook(self.file_path)
        
        sheet_info = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            sheet_info.append({
                "name": sheet_name,
                "max_row": max_row,
                "max_column": max_col,
                "dimensions": f"{max_row}x{max_col}"
            })
        
        return {
            "file_path": self.file_path,
            "filename": os.path.basename(self.file_path),
            "file_size": os.path.getsize(self.file_path),
            "sheet_count": len(workbook.sheetnames),
            "sheets": sheet_info
        }
    
    def get_sheet_data(self, sheet_name: str, range_spec: str = None):
        """Get data from a sheet."""
        from openpyxl import load_workbook
        
        workbook = load_workbook(self.file_path)
        
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}")
        
        sheet = workbook[sheet_name]
        
        if range_spec:
            # Get specific range
            cells = sheet[range_spec]
            if hasattr(cells, '__iter__') and not isinstance(cells, str):
                # Multiple cells
                data = []
                for row in cells:
                    if hasattr(row, '__iter__'):
                        data.append([cell.value for cell in row])
                    else:
                        data.append([row.value])
                actual_range = range_spec
            else:
                # Single cell
                data = [[cells.value]]
                actual_range = range_spec
        else:
            # Get all data
            data = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    data.append(list(row))
            actual_range = f"A1:{sheet.max_column_letter}{sheet.max_row}"
        
        return {
            "data": data,
            "actual_range": actual_range,
            "row_count": len(data),
            "col_count": len(data[0]) if data else 0
        }
    
    def update_cell(self, sheet_name: str, cell_address: str, value):
        """Update a single cell."""
        from openpyxl import load_workbook
        
        workbook = load_workbook(self.file_path)
        
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}")
        
        sheet = workbook[sheet_name]
        
        # Get old value
        old_value = sheet[cell_address].value
        
        # Set new value
        sheet[cell_address] = value
        
        # Save workbook
        workbook.save(self.file_path)
        
        return {
            "sheet": sheet_name,
            "cell": cell_address,
            "old_value": old_value,
            "new_value": value,
            "success": True
        }
    
    def update_range(self, sheet_name: str, range_spec: str, values):
        """Update multiple cells in a range."""
        from openpyxl import load_workbook
        
        workbook = load_workbook(self.file_path)
        
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}")
        
        sheet = workbook[sheet_name]
        
        # Update range
        cells_updated = 0
        cells = sheet[range_spec]
        
        if hasattr(cells, '__iter__') and not isinstance(cells, str):
            for i, row in enumerate(cells):
                if i < len(values):
                    if hasattr(row, '__iter__'):
                        for j, cell in enumerate(row):
                            if j < len(values[i]):
                                cell.value = values[i][j]
                                cells_updated += 1
                    else:
                        if len(values[i]) > 0:
                            row.value = values[i][0]
                            cells_updated += 1
        
        # Save workbook
        workbook.save(self.file_path)
        
        return {
            "sheet": sheet_name,
            "range": range_spec,
            "cells_updated": cells_updated,
            "success": True
        }
    
    def add_sheet(self, sheet_name: str):
        """Add a new sheet."""
        from openpyxl import load_workbook
        
        workbook = load_workbook(self.file_path)
        
        if sheet_name in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' already exists")
        
        # Create new sheet
        workbook.create_sheet(sheet_name)
        
        # Save workbook
        workbook.save(self.file_path)
        
        return {
            "sheet_name": sheet_name,
            "total_sheets": len(workbook.sheetnames),
            "success": True
        }


# ============================================================================
# SERVER STARTUP
# ============================================================================

async def main():
    """
    Main entry point for the Excel MCP server.
    
    This function:
    1. Reads the EXCEL_FILE_PATH environment variable
    2. Validates the Excel file exists
    3. Starts the MCP server with stdio communication
    
    The server expects to be started with:
        EXCEL_FILE_PATH="/path/to/file.xlsx" python3 simple_server.py
    
    MCP clients will communicate with this server via stdin/stdout.
    """
    global current_file_path
    
    # Get Excel file path from environment
    current_file_path = os.getenv("EXCEL_FILE_PATH")
    
    if not current_file_path:
        print("❌ Error: EXCEL_FILE_PATH environment variable not set", file=sys.stderr)
        print("This server requires a specific Excel file path", file=sys.stderr)
        exit(1)
    
    if not os.path.exists(current_file_path):
        print(f"❌ Error: Excel file not found: {current_file_path}", file=sys.stderr)
        exit(1)
    
    print(f"📊 Excel MCP Server (Single File Mode) initialized", file=sys.stderr)
    print(f"📁 Connected file: {os.path.basename(current_file_path)}", file=sys.stderr)
    print(f"🗂️ Full path: {current_file_path}", file=sys.stderr)
    
    # Run the server
    async with mcp.server.stdio.stdio_server() as (read_stream, write_stream):
        await server.run(
            read_stream,
            write_stream,
            InitializationOptions(
                server_name="excel-mcp-single",
                server_version="1.0.0",
                capabilities=types.ServerCapabilities()
            )
        )


# ============================================================================
# SCRIPT ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    """
    Start the MCP server when this script is run directly.
    
    This is typically called by:
    1. The Excel monitor when a user chooses to connect a file
    2. Manual testing: EXCEL_FILE_PATH="file.xlsx" python3 simple_server.py
    3. Claude Desktop configuration in claude_desktop_config.json
    """
    asyncio.run(main())